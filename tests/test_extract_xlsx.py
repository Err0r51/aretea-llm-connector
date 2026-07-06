"""AC3: extract_xlsx returns cell values, never formula strings (real .xlsx)."""

from __future__ import annotations

import io

from aretea_drive_mcp.gdrive.extract import extract_xlsx
from openpyxl import Workbook

CEIL = 314_572_800


def _xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    ws["A1"] = "revenue"
    ws["B1"] = 1234
    ws["A2"] = "=1+1"  # a formula: with no cached value, data_only reads blank — never the string
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_xlsx_returns_values_not_formulas() -> None:
    text = extract_xlsx(_xlsx_bytes(), max_uncompressed_bytes=CEIL)
    assert "# Budget" in text
    assert "revenue" in text
    assert "1234" in text
    assert "=1+1" not in text  # formula strings must never leak into output
